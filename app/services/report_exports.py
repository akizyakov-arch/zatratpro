from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from app.ui.reports import REPORT_PERIOD_LABELS


def build_manager_report_workbook(
    period: str,
    summary,
    projects,
    employees,
    duplicates,
    documents,
    items,
) -> tuple[str, bytes]:
    workbook = Workbook()

    overview = workbook.active
    overview.title = 'Overview'
    overview.append(['Период', REPORT_PERIOD_LABELS.get(period, period)])
    overview.append(['С даты', summary.start_at.isoformat()])
    overview.append(['Документов', summary.documents])
    overview.append(['Сумма', float(summary.total_amount or 0)])
    overview.append(['Точных дублей', summary.exact_duplicates])
    overview.append(['Вероятных дублей', summary.probable_duplicates])

    projects_sheet = workbook.create_sheet('Projects')
    projects_sheet.append(['ID проекта', 'Проект', 'Документов', 'Сумма', 'Точных дублей', 'Вероятных дублей'])
    for row in projects:
        projects_sheet.append([row.project_id, row.project_name, row.document_count, float(row.total_amount or 0), row.exact_duplicate_count, row.probable_duplicate_count])

    employees_sheet = workbook.create_sheet('Employees')
    employees_sheet.append(['ID пользователя', 'Сотрудник', 'Username', 'Документов', 'Сумма', 'Точных дублей', 'Вероятных дублей'])
    for row in employees:
        employees_sheet.append([row.user_id, row.employee_name or '', row.username or '', row.document_count, float(row.total_amount or 0), row.exact_duplicate_count, row.probable_duplicate_count])

    duplicates_sheet = workbook.create_sheet('Duplicates')
    duplicates_sheet.append(['ID документа', 'Статус дубля', 'Дубль записи', 'Проект', 'Номер', 'Дата', 'Продавец', 'ИНН', 'Сумма', 'Кем загружен'])
    for row in duplicates:
        duplicates_sheet.append([
            row.id,
            row.duplicate_status,
            row.duplicate_of_document_id or '',
            row.project_name,
            row.document_number or '',
            row.document_date.isoformat() if row.document_date else '',
            row.vendor or '',
            row.vendor_inn or '',
            float(row.total_amount or 0),
            row.uploaded_by_name or '',
        ])

    documents_sheet = workbook.create_sheet('Documents')
    documents_sheet.append(['ID', 'Проект', 'Номер', 'Дата', 'Продавец', 'ИНН', 'Сумма', 'Статус дубля', 'Загружен', 'Кем загружен'])
    for row in documents:
        documents_sheet.append([
            row.id,
            row.project_name,
            row.document_number or '',
            row.document_date.isoformat() if row.document_date else '',
            row.vendor or '',
            row.vendor_inn or '',
            float(row.total_amount or 0),
            row.duplicate_status,
            row.created_at.isoformat(),
            row.uploaded_by_name or '',
        ])

    items_sheet = workbook.create_sheet('Items')
    items_sheet.append(['ID документа', 'Строка', 'Наименование', 'Количество', 'Цена', 'Сумма'])
    for row in items:
        items_sheet.append([row.document_id, row.line_no, row.name or '', float(row.quantity or 0), float(row.price or 0), float(row.line_total or 0)])

    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = '' if cell.value is None else str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 40)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return f'reports-{period}.xlsx', buffer.getvalue()
